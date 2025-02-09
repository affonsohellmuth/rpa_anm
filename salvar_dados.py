from openpyxl import load_workbook, Workbook
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from utils import data_hora_atual
from multiprocessing import Lock
import traceback
import pandas as pd
import os
import time

nome_arquivo = f"Maiores_Arrecadadores_{data_hora_atual}.xlsx"
pasta_planilha = os.path.join(os.path.expanduser("~"), "Documents/Planilha ANM")

def capturar_primeiras_seis_colunas(func_navegador):
    try:
        elemento = func_navegador.find_element(By.CSS_SELECTOR, '#ctl00_ContentPlaceHolder1_dvResultado > table.tabelaFormulario')
        texto = elemento.text
    except NoSuchElementException as e:
        print(f"❌ Erro ao capturar elemento: {e}")
        return {}
    except TimeoutException:
        print("❌ Timeout - A página não carregou a tempo. Tentando atualizar...")
        func_navegador.refresh()  # Tenta atualizar a página
        time.sleep(5)  # Aguarda 5 segundos para a página recarregar
        return capturar_primeiras_seis_colunas(func_navegador)  # Tenta novamente após o refresh

    dados = {}
    for linha in texto.split("\n"):  
        if " : " in linha:  
            chave, valor = linha.split(" : ", 1)  
            dados[chave.strip()] = valor.strip()  

    #print(dados)
    return dados  

def salvar_dados_planilha(dados_tabela, nome_arquivo):
    if not os.path.exists(pasta_planilha):
        os.makedirs(pasta_planilha)
        print(f"📂 Pasta 'Planilha ANM' criada em: {pasta_planilha}")
    
    caminho_arquivo = os.path.join(pasta_planilha, nome_arquivo)

    if os.path.exists(caminho_arquivo):
        workbook = load_workbook(caminho_arquivo)
    else:
        workbook = Workbook()

    folha = workbook.active
    folha.title = "Arrecadadores"

    if folha.max_row == 1 and folha.cell(row=1, column=1).value != "Arrecadador (Empresa)":
        folha.append(["Arrecadador (Empresa)", "Qtde Títulos", "Operação", "Recolhimento CFEM", "% Recolhimento CFEM"])

    for dado in dados_tabela:
        folha.append([  
            dado["Arrecadador (Empresa)"],
            dado["Qtde Títulos"],
            dado["Operação"],
            dado["Recolhimento CFEM"],
            dado["% Recolhimento CFEM"]
        ])

    workbook.save(caminho_arquivo)
    print(f"✅ Dados salvos na planilha '{caminho_arquivo}' com sucesso!")

def capturar_todos_os_dados(func_navegador, municipio, regiao):
    try:
        # Captura os dados da primeira tabela
        elemento = func_navegador.find_element(By.CSS_SELECTOR, '#ctl00_ContentPlaceHolder1_dvResultado > table.tabelaFormulario')
        texto = elemento.text
    except NoSuchElementException as e:
        print(f"❌ Erro ao capturar dados da primeira tabela: {e}")
        return []  # Retorna uma lista vazia em caso de erro
    except TimeoutException:
        print("❌ Timeout - A página não carregou a tempo. Tentando atualizar...")
        func_navegador.refresh()  # Tenta atualizar a página
        time.sleep(5)  # Aguarda 5 segundos para a página recarregar
        return capturar_todos_os_dados(func_navegador, municipio, regiao)  # Tenta novamente após o refresh

    dados_gerais = {}
    for linha in texto.split("\n"):
        if " : " in linha:
            chave, valor = linha.split(" : ", 1)
            dados_gerais[chave.strip()] = valor.strip()

    dados_gerais["Região"] = regiao
    #print("Dados gerais capturados:", dados_gerais)

    try:
        # Captura os dados da segunda tabela
        tabela = func_navegador.find_element(By.CSS_SELECTOR, '#ctl00_ContentPlaceHolder1_dvResultado > table.tabelaRelatorio')
        linhas = tabela.find_elements(By.CSS_SELECTOR, "tbody > tr")
    except NoSuchElementException as e:
        print(f"❌ Erro ao capturar a segunda tabela: {e}")
        return []  # Retorna uma lista vazia em caso de erro
    except TimeoutException:
        print("❌ Timeout - A página não carregou a tempo. Tentando atualizar...")
        func_navegador.refresh()  # Tenta atualizar a página
        time.sleep(5)  # Aguarda 5 segundos para a página recarregar
        return capturar_todos_os_dados(func_navegador, municipio, regiao)

    dados_completos = []
    
    if not linhas:  
        #print("Tabela vazia. Preenchendo com dados gerais e valores zerados.")
        dados_completos.append({
            "Arrecadador (Empresa)": "",
            "Qtde Títulos": 0,
            "Operação": "",
            "Recolhimento CFEM": 0,
            "% Recolhimento CFEM": 0,
            **dados_gerais  
        })
    else:  
        for linha in linhas:
            colunas = linha.find_elements(By.TAG_NAME, "td")
            if len(colunas) >= 6:
                linha_dados = {
                    "Arrecadador (Empresa)": colunas[1].text,
                    "Qtde Títulos": colunas[2].text,
                    "Operação": colunas[3].text,
                    "Recolhimento CFEM": colunas[4].text,
                    "% Recolhimento CFEM": colunas[5].text,
                    **dados_gerais
                }
                dados_completos.append(linha_dados)
            else:
                print(f"❌ Linha ignorada por não ter colunas suficientes: {linha.text}")

    for dado in dados_completos:
        print("✅Dados completos capturados")

    return dados_completos


# Configurações
PASTA_TEMP = os.path.join(os.getcwd(), "temp_arrecadadores")
os.makedirs(PASTA_TEMP, exist_ok=True)
lock = Lock()


def salvar_dados_completos_planilha(dados_completos, processo_id):
    pasta_planilha = os.path.join(os.path.expanduser("~"), "Documents/Planilha ANM")
    os.makedirs(pasta_planilha, exist_ok=True)

    # Nome do arquivo temporário ÚNICO por processo (sem timestamp)
    nome_temp = f"temp_processo_{processo_id}.xlsx"
    caminho_temp = os.path.join(PASTA_TEMP, nome_temp)

    # Converter dados para DataFrame
    df_novo = pd.DataFrame(dados_completos)
    df_novo = df_novo.loc[:, ~df_novo.columns.duplicated()]

    try:
        print(f"⏳ Salvando dados temporariamente em: {caminho_temp}")

        with lock:  # 🔒 Evita concorrência entre processos
            if os.path.exists(caminho_temp):
                # Lê o arquivo para descobrir a última linha preenchida
                df_existente = pd.read_excel(caminho_temp, engine='openpyxl')
                with pd.ExcelWriter(caminho_temp, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                    df_novo.to_excel(writer, index=False, sheet_name='Sheet1', header=False,
                                     startrow=len(df_existente) + 1)
            else:
                # Criar um novo arquivo com o cabeçalho
                with pd.ExcelWriter(caminho_temp, mode='w', engine='openpyxl') as writer:
                    df_novo.to_excel(writer, index=False, sheet_name='Sheet1')

    except Exception as e:
        print(f"❌ Erro ao salvar dados (processo {processo_id}): {e}")

def verificar_arquivo_excel(caminho_arquivo):
    try:
        pd.read_excel(caminho_arquivo, engine='openpyxl')  # Tenta abrir
        return True
    except Exception:
        return False  # Se der erro, significa que o arquivo está corrompido ou incompleto

def unificar_planilhas(nome_arquivo_final):
    try:
        # Caminho do arquivo final
        pasta_final = os.path.join(os.path.expanduser("~"), "Documents/Planilha ANM")
        os.makedirs(pasta_final, exist_ok=True)
        caminho_final = os.path.join(pasta_final, nome_arquivo_final)

        # Listar arquivos temporários
        arquivos_temp = [
            os.path.join(PASTA_TEMP, f)
            for f in os.listdir(PASTA_TEMP)
            if f.startswith("temp_processo") and f.endswith(".xlsx") and verificar_arquivo_excel(
                os.path.join(PASTA_TEMP, f))
        ]

        if not arquivos_temp:
            print("❌ Nenhum arquivo temporário encontrado para unificação.")
            return

        # Combinar dados
        dfs = []
        for arquivo in arquivos_temp:
            df = pd.read_excel(arquivo, engine='openpyxl')
            dfs.append(df)

        df_final = pd.concat(dfs, ignore_index=True)

        # Salvar arquivo final
        with lock:
            df_final.to_excel(caminho_final, index=False, engine='openpyxl')

        print(f"✅ Planilha final salva em: {caminho_final}")

        # Limpar arquivos temporários (opcional)
        for arquivo in arquivos_temp:
            os.remove(arquivo)

        # Remover a pasta apenas se estiver vazia
        if os.path.exists(PASTA_TEMP) and not os.listdir(PASTA_TEMP):
            os.rmdir(PASTA_TEMP)

        print("✅ Arquivos temporários removidos.")

    except Exception as e:
        print(f"❌  Erro ao unificar planilhas: {e}")
        traceback.print_exc()

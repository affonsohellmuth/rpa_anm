from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
import pandas as pd
import time
import os

# Configuração inicial do Selenium
driver = webdriver.Chrome()  # Substitua pelo seu driver adequado
driver.maximize_window()
url = "https://sistemas.anm.gov.br/arrecadacao/extra/relatorios/cfem/maiores_arrecadadores.aspx"
driver.get(url)

# Função para esperar o carregamento
def esperar_carregamento():
    time.sleep(2)  # Ajuste conforme necessário

# Função para clicar em um elemento
def clicar_elemento(xpath):
    driver.find_element(By.XPATH, xpath).click()
    esperar_carregamento()

# Função para selecionar uma opção em um combo box
def selecionar_combo(xpath, valor):
    select = Select(driver.find_element(By.XPATH, xpath))
    select.select_by_visible_text(valor)
    esperar_carregamento()

# Função para extrair os dados da tabela
def extrair_dados():
    try:
        tabela = driver.find_element(By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_dvResultado"]/table[2]')
        linhas = tabela.find_elements(By.TAG_NAME, 'tr')
        dados = []

        for linha in linhas[1:]:  # Ignorar a primeira linha (cabeçalho)
            colunas = linha.find_elements(By.TAG_NAME, 'td')
            # Ignorar a linha que contém "TOTAL"
            if colunas[0].text.strip().upper() == "Total":
                continue
            dados.append([coluna.text.strip() for coluna in colunas])

        return dados if dados else None
    except Exception as e:
        print("Erro ao extrair dados:", e)
        return None

# Função para pegar as opções de um combo box
def obter_opcoes(xpath):
    select = Select(driver.find_element(By.XPATH, xpath))
    return [opcao.text for opcao in select.options]

# Função para gravar os dados no Excel sem sobrescrever
def gravar_dados(df, arquivo='maiores_arrecadadores_completo.xlsx'):
    if os.path.exists(arquivo):
        # Abrir o arquivo existente e adicionar dados no final
        with pd.ExcelWriter(arquivo, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Descobrir a última linha
            workbook = load_workbook(arquivo)
            sheet = workbook.active
            ultima_linha = sheet.max_row
            df.to_excel(writer, index=False, header=False, startrow=ultima_linha)
    else:
        # Criar o arquivo com cabeçalho
        df.to_excel(arquivo, index=False)

# Função principal para processar os combos e extrair os dados
def processar():
    selecionar_combo('//*[@id="ctl00_ContentPlaceHolder1_nu_Ano"]', '2025')
    selecionar_combo('//*[@id="ctl00_ContentPlaceHolder1_regiao"]', 'Centro-Oeste')

    arrecadado_por = driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_rdComparacao_5")
    ordenado_por = driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_rdOrdenacao_0")

    if not arrecadado_por.is_selected():
        arrecadado_por.click()
    
    if not ordenado_por.is_selected():
        ordenado_por.click()

    esperar_carregamento()

    subs_agrupadoras = obter_opcoes('//*[@id="ctl00_ContentPlaceHolder1_subs_agrupadora"]')
    for sub_agrupadora in subs_agrupadoras:
        if sub_agrupadora == "Todas as Agrupadoras":
            continue
        selecionar_combo('//*[@id="ctl00_ContentPlaceHolder1_subs_agrupadora"]', sub_agrupadora)

        substancias = obter_opcoes('//*[@id="ctl00_ContentPlaceHolder1_Substancia"]')
        for substancia in substancias:
            if substancia == "Todas as Substância":
                continue
            selecionar_combo('//*[@id="ctl00_ContentPlaceHolder1_Substancia"]', substancia)

            estados = obter_opcoes('//*[@id="ctl00_ContentPlaceHolder1_Estado"]')
            for estado in estados:
                if estado == "Todos os Estados":
                    continue
                selecionar_combo('//*[@id="ctl00_ContentPlaceHolder1_Estado"]', estado)

                municipios = obter_opcoes('//*[@id="ctl00_ContentPlaceHolder1_Municipio"]')
                for municipio in municipios:
                    if municipio in ["Todas os Município", "***"]:
                        continue
                    selecionar_combo('//*[@id="ctl00_ContentPlaceHolder1_Municipio"]', municipio)

                    clicar_elemento('//*[@id="ctl00_ContentPlaceHolder1_btnGera"]')

                    dados = extrair_dados()

                    if dados:
                        df = pd.DataFrame(dados, columns=[
                            'Arrecadador (Empresa)', 'Qtde Títulos', 'Operação', 'Recolhimento CFEM'
                        ])
                        df['% Recolhimento CFEM'] = '0,00%'
                    else:
                        print(f"Tabela vazia para Município: {municipio}.")
                        df = pd.DataFrame(columns=[
                            'Arrecadador (Empresa)', 'Qtde Títulos', 'Operação', 'Recolhimento CFEM', '% Recolhimento CFEM'
                        ])

                    # Adicionar contexto
                    df.insert(0, 'Município', municipio)
                    df.insert(0, 'Estado', estado)
                    df.insert(0, 'Região', 'Centro-Oeste')
                    df.insert(0, 'Substância', substancia)
                    df.insert(0, 'Subs.Agrupadora', sub_agrupadora)
                    df.insert(0, 'Ano', '2025')

                    # Gravar no Excel
                    gravar_dados(df)
                    print(f"Dados do município {municipio} gravados com sucesso.")

# Iniciar o processo
processar()

# Fechar o driver ao final
driver.quit()
